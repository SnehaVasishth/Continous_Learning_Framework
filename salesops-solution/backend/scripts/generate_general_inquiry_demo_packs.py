"""Generate two general-inquiry demo packs in ``~/Downloads/`` that reference
Salesforce Orders we seeded for Leeway Hertz. The Stage 5 reply drafter has
been wired to consume ``customer_match.salesforce.recent_orders`` directly,
so when these emails land in the platform the generated reply quotes the
authoritative ship dates from Salesforce instead of echoing the customer's
own message back.

Seeded Salesforce Orders on the Leeway Hertz account:
  * PO PO-LH-Q3-1814 -> EndDate 2026-07-22 (N9020A MXA + Option B25, Pune Lab A)
  * PO PO-LH-Q3-1817 -> EndDate 2026-07-24 (N9020A MXA + Option B25, Pune Lab B)
  * PO PO-LH-Q3-1823 -> EndDate 2026-08-05 (N9020A MXA + Option B25, Pune Lab B)
  * PO PO-LH-Q2-1714 -> EndDate 2026-05-30 (E36313B + 34461A, Pune Power Lab)
  * PO PO-LH-Q3-2121 -> EndDate 2026-09-01 (N5230C PNA-L 13.5 GHz, Hyderabad RF)

Two demo emails are generated:
  1. Lead-time confirmation across the three N9020A orders (1814, 1817, 1823).
     Customer asks for confirmation against a 25 July ask. The reply will
     quote 22-Jul and 24-Jul (on time) and 5-Aug (at risk) from Salesforce.
  2. Quarterly open-order review across all five POs. Customer asks for the
     current status and committed dates. The reply will quote the live
     Salesforce status (Activated / Draft) and EndDate per PO.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer


OUTPUT_ROOT = Path.home() / "Downloads" / "leewayhertz-general-inquiry-packs"


DEMOS = [
    {
        "slug": "01-lead-time-three-n9020a-orders",
        "intent_hint": "general_inquiry",
        "subject": "Lead-time confirmation for three open N9020A orders (PO-LH-Q3-1814, 1817, 1823)",
        "body": (
            "Hello Keysight Trade Order team,\n\n"
            "We are finalising our Q3 production schedule and need confirmation on the "
            "committed ship dates for three N9020A MXA orders we have open with you. "
            "Our internal stage gate closes on 18 June, and our Pune lab needs all three "
            "units on dock by 25 July at the latest so the audited acceptance window can begin.\n\n"
            "The open POs on our side are:\n"
            "  - PO-LH-Q3-1814 (Pune Lab A)\n"
            "  - PO-LH-Q3-1817 (Pune Lab B)\n"
            "  - PO-LH-Q3-1823 (Pune Lab B)\n\n"
            "The attached worksheet lists the three POs with the line numbers and the dates "
            "we have on our side. The PDF is our internal standard N9020A configuration "
            "(Option B25 wideband demod is mandatory across the lab rollout).\n\n"
            "Please confirm the current Keysight-side EndDate (committed ship date) for each "
            "of the three orders, the build status for the Option B25 line where applicable, "
            "and whether the three units can ship consolidated from the Singapore hub.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "LH_Open_PO_lead_time_tracker.xlsx",
            "sheet_name": "Q3-2026 N9020A POs",
            "title": "Leeway Hertz - Open PO lead-time tracker (Q3 2026)",
            "header": ["PO Number", "Line", "Part Number", "Description", "Qty", "Customer-Requested Delivery", "Ship-To", "Notes"],
            "rows": [
                ["PO-LH-Q3-1814", 1, "N9020A-MXA-3.6", "MXA Signal Analyzer, 3.6 GHz preset", 1, "2026-07-21", "Pune Lab A", "B25 required"],
                ["PO-LH-Q3-1814", 2, "N9020A-OPT-B25", "Wideband Digital Demodulation", 1, "2026-07-21", "Pune Lab A", "Build-to-order"],
                ["PO-LH-Q3-1817", 1, "N9020A-MXA-3.6", "MXA Signal Analyzer, 3.6 GHz preset", 1, "2026-07-25", "Pune Lab B", "B25 required"],
                ["PO-LH-Q3-1817", 2, "N9020A-OPT-B25", "Wideband Digital Demodulation", 1, "2026-07-25", "Pune Lab B", "Ship with line 1"],
                ["PO-LH-Q3-1823", 1, "N9020A-MXA-3.6", "MXA Signal Analyzer, 3.6 GHz preset", 1, "2026-07-25", "Pune Lab B", "B25 required"],
                ["PO-LH-Q3-1823", 2, "N9020A-OPT-B25", "Wideband Digital Demodulation", 1, "2026-07-25", "Pune Lab B", "Consolidate freight OK"],
            ],
            "footer": (
                "Stage gate closes 2026-06-18. Customer needs all three units on dock at Pune by 2026-07-25. "
                "Consolidated freight from Singapore hub preferred over split shipments."
            ),
        },
        "pdf": {
            "filename": "LH_N9020A_standard_config.pdf",
            "title": "Leeway Hertz - Standard N9020A MXA Configuration (Q3 2026 rollout)",
            "intent": "Internal procurement standard for all new N9020A units acquired in 2026.",
            "sections": [
                ("Base instrument", [
                    "Model: Keysight N9020A MXA Signal Analyzer",
                    "Frequency range: 20 Hz to 3.6 GHz (preset)",
                    "Resolution bandwidth: 1 Hz to 8 MHz",
                ]),
                ("Mandatory options", [
                    "Option B25 - Wideband Digital Demodulation (mandatory across all new units)",
                    "Option EXM - External mixing capability",
                    "3-year calibration contract (CalSure Gold)",
                ]),
                ("Sites in scope", [
                    "Pune Lab A - N9020A x 1 (Q3 2026)",
                    "Pune Lab B - N9020A x 2 (Q3 2026)",
                ]),
                ("Acceptance gate", [
                    "Unit must arrive with current Keysight calibration certificate, valid 12 months from ship date.",
                    "Option B25 must be factory-installed; field-install is not accepted for this rollout.",
                ]),
            ],
            "footer": "Document owner: Procurement Lead, Leeway Hertz. Contact: rituraj@leewayhertz.com.",
        },
    },
    {
        "slug": "02-quarterly-open-order-review",
        "intent_hint": "general_inquiry",
        "subject": "Quarterly open-order review - current status across five Leeway Hertz POs",
        "body": (
            "Hi Keysight Sales Operations,\n\n"
            "We are running our quarterly open-order review and need an authoritative "
            "status read across the five POs Leeway Hertz currently has open with you. "
            "Our finance team needs the committed ship dates and current order status to "
            "finalise the Q3 accrual.\n\n"
            "Open POs on our side:\n"
            "  - PO-LH-Q2-1714 (Pune Power Lab) - E36313B + 34461A\n"
            "  - PO-LH-Q3-1814 (Pune Lab A) - N9020A + B25\n"
            "  - PO-LH-Q3-1817 (Pune Lab B) - N9020A + B25\n"
            "  - PO-LH-Q3-1823 (Pune Lab B) - N9020A + B25\n"
            "  - PO-LH-Q3-2121 (Hyderabad RF Lab) - N5230C + 85052D\n\n"
            "The attached worksheet is our internal open-order register with the latest "
            "values we have on our side. The PDF is the section of our master agreement "
            "that defines our standard quarterly-review protocol so the review aligns with "
            "the contractual cadence.\n\n"
            "Please reply with the current Keysight-side OrderNumber, Status, and committed "
            "EndDate for each PO. Flag any order where the committed date has slipped from "
            "what we have on record so we can flag the variance to finance.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "LH_Open_Order_Register_Q3-2026.xlsx",
            "sheet_name": "Open orders register",
            "title": "Leeway Hertz - Open order register (as of quarterly review)",
            "header": ["PO Number", "Lab / Ship-To", "Product Family", "Our Expected Ship Date", "Status (LH side)", "Owner (LH)", "Notes"],
            "rows": [
                ["PO-LH-Q2-1714", "Pune Power Lab", "E36313B + 34461A", "2026-05-30", "Awaiting Keysight ack", "Anika Mehta", "Finance accrual depends on this date"],
                ["PO-LH-Q3-1814", "Pune Lab A", "N9020A + Opt B25", "2026-07-21", "Awaiting Keysight ack", "Vikram Kapoor", "Stage-gate dependency"],
                ["PO-LH-Q3-1817", "Pune Lab B", "N9020A + Opt B25", "2026-07-25", "Awaiting Keysight ack", "Vikram Kapoor", "Stage-gate dependency"],
                ["PO-LH-Q3-1823", "Pune Lab B", "N9020A + Opt B25", "2026-07-25", "Awaiting Keysight ack", "Vikram Kapoor", "B25 component sensitive"],
                ["PO-LH-Q3-2121", "Hyderabad RF Lab", "N5230C + 85052D", "2026-09-01", "Awaiting Keysight ack", "Sanjay Iyer", "Q3 capex closeout"],
            ],
            "footer": (
                "Five open Keysight POs on the Leeway Hertz account as of the quarterly review. "
                "Variance from the Keysight-side dates needs to be flagged to finance before quarter-end."
            ),
        },
        "pdf": {
            "filename": "LH_MasterAgreement_QuarterlyReview_Section.pdf",
            "title": "Leeway Hertz - Master Agreement, Quarterly Review Section (extract)",
            "intent": "Extract from Master Agreement MA-LH-KS-2025-04-01 covering Section 11 (open-order review).",
            "sections": [
                ("Section 11.1 - Cadence", [
                    "Joint open-order review held in the final week of each calendar quarter.",
                    "Keysight provides an authoritative read of OrderNumber, Status, and committed EndDate per open PO.",
                    "Leeway Hertz reconciles against the internal open-order register and flags any variance.",
                ]),
                ("Section 11.2 - Required fields", [
                    "Salesforce OrderNumber",
                    "Customer PO Number",
                    "Status (Draft, Activated, Cancelled)",
                    "EffectiveDate and EndDate",
                    "Ship-to facility",
                ]),
                ("Section 11.3 - Variance handling", [
                    "Date variance over 5 business days triggers a written flag to procurement and finance.",
                    "Variance is reconciled within 10 business days; resolution captured on the Salesforce Case.",
                ]),
            ],
            "footer": "Reference: MA-LH-KS-2025-04-01, Section 11 extract. Procurement contact: rituraj@leewayhertz.com.",
        },
    },
]


def write_email_txt(folder: Path, demo: dict) -> Path:
    path = folder / "email.txt"
    body = (
        f"FROM:    rituraj@leewayhertz.com\n"
        f"TO:      [Keysight sandbox mailbox the demo instance is polling]\n"
        f"SUBJECT: {demo['subject']}\n\n"
        f"{demo['body']}\n"
    )
    path.write_text(body, encoding="utf-8")
    return path


def write_excel(folder: Path, demo: dict) -> Path:
    spec = demo["excel"]
    path = folder / spec["filename"]
    wb = Workbook()
    ws = wb.active
    ws.title = spec["sheet_name"][:31]
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Side(border_style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=1, value=spec["title"]).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(spec["header"]))

    header_row = 3
    for col_idx, h in enumerate(spec["header"], start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    for r_idx, row in enumerate(spec["rows"], start=header_row + 1):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = border
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")

    footer_row = header_row + 1 + len(spec["rows"]) + 1
    ws.cell(row=footer_row, column=1, value=spec.get("footer", "")).font = Font(italic=True, color="6B7280")
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(spec["header"]))

    for col_idx in range(1, len(spec["header"]) + 1):
        letter = ws.cell(row=header_row, column=col_idx).column_letter
        max_len = len(str(spec["header"][col_idx - 1]))
        for row in spec["rows"]:
            v = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 38)

    wb.save(path)
    return path


def write_pdf(folder: Path, demo: dict) -> Path:
    spec = demo["pdf"]
    path = folder / spec["filename"]
    doc = SimpleDocTemplate(
        str(path),
        pagesize=LETTER,
        leftMargin=0.85 * inch,
        rightMargin=0.85 * inch,
        topMargin=0.85 * inch,
        bottomMargin=0.85 * inch,
        title=spec["title"],
        author="Leeway Hertz Procurement",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14, leading=18, spaceAfter=10, textColor=colors.HexColor("#1F4E79"))
    intent = ParagraphStyle("Intent", parent=styles["Italic"], fontSize=9.5, leading=13, textColor=colors.HexColor("#6B7280"), spaceAfter=14)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=11.5, leading=16, spaceBefore=10, spaceAfter=4, textColor=colors.HexColor("#111827"))
    body = ParagraphStyle("Body", parent=styles["BodyText"], fontSize=10.5, leading=15, spaceAfter=3, textColor=colors.HexColor("#111827"))
    footer = ParagraphStyle("Footer", parent=styles["Italic"], fontSize=9, leading=12, textColor=colors.HexColor("#6B7280"), spaceBefore=18)

    story = []
    story.append(Paragraph(spec["title"], h1))
    if spec.get("intent"):
        story.append(Paragraph(spec["intent"], intent))
    for section_title, bullets in spec["sections"]:
        story.append(Paragraph(section_title, h2))
        for b in bullets:
            story.append(Paragraph(f"&bull; {b}", body))
    if spec.get("footer"):
        story.append(Paragraph(spec["footer"], footer))
    doc.build(story)
    return path


def write_readme(folder: Path, demo: dict) -> Path:
    path = folder / "README.txt"
    excel_name = demo["excel"]["filename"]
    pdf_name = demo["pdf"]["filename"]
    body = (
        f"Demo pack: {demo['slug']}\n"
        f"Expected intent: general_inquiry\n"
        f"{'=' * 60}\n\n"
        f"This case demonstrates a Stage 5 reply grounded in Salesforce data.\n"
        f"The platform pulls the open Leeway Hertz Orders from the ZBrain\n"
        f"sandbox environment at Stage 2.4 (customer enrichment), passes them\n"
        f"into the Stage 5 reply prompt as SALESFORCE CONTEXT, and the LLM\n"
        f"quotes the committed EndDate per PO directly from those rows.\n\n"
        f"How to run this case:\n"
        f"  1. From your email client signed in as rituraj@leewayhertz.com,\n"
        f"     compose a new message to the demo's polling mailbox.\n"
        f"  2. Paste the subject and body from email.txt below.\n"
        f"  3. Attach BOTH files in this folder:\n"
        f"       - {excel_name}\n"
        f"       - {pdf_name}\n"
        f"  4. Send the email. IMAP picks it up within ten seconds.\n"
        f"  5. Watch the Trace page; the Stage 5 reply should quote the\n"
        f"     Keysight-side EndDate per PO (pulled live from Salesforce),\n"
        f"     not the date the customer asked about in their email.\n\n"
        f"Subject (copy verbatim):\n"
        f"  {demo['subject']}\n"
    )
    path.write_text(body, encoding="utf-8")
    return path


def main() -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    print(f"Writing demo packs under {OUTPUT_ROOT}")
    for demo in DEMOS:
        folder = OUTPUT_ROOT / demo["slug"]
        folder.mkdir(parents=True, exist_ok=True)
        write_email_txt(folder, demo)
        write_excel(folder, demo)
        write_pdf(folder, demo)
        write_readme(folder, demo)
        print(f"  [{demo['slug']}] email.txt + {demo['excel']['filename']} + {demo['pdf']['filename']}")
    index = OUTPUT_ROOT / "INDEX.txt"
    lines = [
        "Leeway Hertz general-inquiry demo packs",
        "=" * 60,
        "",
        "These two cases exercise the Stage 5 reply grounding fix:",
        "the LLM now quotes Salesforce Order data (EndDate, Status)",
        "directly from customer_match.salesforce.recent_orders, rather",
        "than echoing dates the customer mentions in their email.",
        "",
        "Salesforce sandbox setup (already provisioned):",
        "  Contact: rituraj@leewayhertz.com under Leeway Hertz Account",
        "  Orders seeded on the Leeway Hertz Account:",
        "    PO-LH-Q3-1814  EndDate 2026-07-22  Draft      N9020A+B25  Pune Lab A",
        "    PO-LH-Q3-1817  EndDate 2026-07-24  Draft      N9020A+B25  Pune Lab B",
        "    PO-LH-Q3-1823  EndDate 2026-08-05  Draft      N9020A+B25  Pune Lab B",
        "    PO-LH-Q2-1714  EndDate 2026-05-30  Draft      E36313B+DMM Pune Power",
        "    PO-LH-Q3-2121  EndDate 2026-09-01  Draft      N5230C+kit  Hyderabad RF",
        "",
        f"{len(DEMOS)} demo cases:",
    ]
    for d in DEMOS:
        lines.append(f"  - {d['slug']:34s} {d['subject']}")
    lines.append("")
    lines.append("Each subfolder contains email.txt + Excel + PDF + README.txt.")
    index.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {index}")


if __name__ == "__main__":
    main()
