"""Generate two hold-release demo packs under
``~/Downloads/leewayhertz-hold-release-packs/``.

These reference two Leeway Hertz orders that are tagged as on-hold in
Salesforce (OrderReferenceNumber LIKE 'HOLD-%'):

  * PO-LH-Q3-3014 (OrderNumber 00000138, HOLD-OVERDUE-INVOICE,
    N9030B PXA Signal Analyzer 50 GHz with Option B85, Pune RF Lab,
    EndDate 2026-07-18)
  * PO-LH-Q3-3017 (OrderNumber 00000139, HOLD-CREDIT-LIMIT,
    M9421A VXT Vector Transceiver bundle + 85052D cal kit,
    Hyderabad RF Lab, EndDate 2026-07-29)

Stage 2.4's hold_release enrichment fetches these as orders_on_hold, and
the Stage 5 reply prompt grounds the customer reply on the SF Order data.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate


OUTPUT_ROOT = Path.home() / "Downloads" / "leewayhertz-hold-release-packs"


DEMOS = [
    {
        "slug": "01-hold-release-PO-LH-Q3-3014-overdue-invoice",
        "sender": "rituraj@leewayhertz.com",
        "subject": "Hold release on PO-LH-Q3-3014 - overdue invoices fully settled, please release",
        "body": (
            "Hello Keysight Order Operations,\n\n"
            "Our finance team confirmed this morning that the two overdue invoices flagged "
            "against PO-LH-Q3-3014 (N9030B PXA Signal Analyzer 50 GHz with Option B85, "
            "ship-to Pune RF Lab) have cleared on the wire we sent yesterday. "
            "Our treasury team's wire reference is WIRE-LH-2026-06-2104, "
            "value date 2026-06-12, total settled USD 21,485 against invoices "
            "INV-KS-2026-05-7102 and INV-KS-2026-05-7118.\n\n"
            "Please lift the credit hold on PO-LH-Q3-3014 today. Our Pune RF Lab is "
            "queuing the calibration acceptance window the day after the unit lands, "
            "and any further hold will slip the lab acceptance schedule.\n\n"
            "Attached are the wire confirmation and the treasury payment-proof worksheet.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "Payment_proof_PO-LH-Q3-3014.xlsx",
            "sheet_name": "Wire payment ledger",
            "title": "Leeway Hertz - Payment proof for PO-LH-Q3-3014",
            "header": ["Keysight Invoice", "Our PO Reference", "Amount (USD)", "Bank Wire Reference", "Value Date", "Status"],
            "rows": [
                ["INV-KS-2026-05-7102", "PO-LH-Q3-3014", "18,950.00", "WIRE-LH-2026-06-2104", "2026-06-12", "Cleared"],
                ["INV-KS-2026-05-7118", "PO-LH-Q3-3014",  "2,535.00", "WIRE-LH-2026-06-2104", "2026-06-12", "Cleared"],
                ["TOTAL", "PO-LH-Q3-3014", "21,485.00", "WIRE-LH-2026-06-2104", "2026-06-12", "Cleared"],
            ],
            "footer": "PO-LH-Q3-3014 invoices fully settled. Please release credit hold today.",
        },
        "pdf": {
            "filename": "SWIFT_confirmation_PO-LH-Q3-3014.pdf",
            "title": "SWIFT MT103 confirmation - Leeway Hertz to Keysight",
            "intent": "Outbound wire confirmation against PO-LH-Q3-3014.",
            "sections": [
                ("Wire instruction", [
                    "Sender: Leeway Hertz Pvt Ltd",
                    "Beneficiary: Keysight Technologies Singapore Pte Ltd",
                    "Amount: USD 21,485.00",
                    "Value date: 2026-06-12",
                    "Customer reference: WIRE-LH-2026-06-2104",
                ]),
                ("Application", [
                    "Apply against Order under PO-LH-Q3-3014 and release credit hold.",
                    "Invoices in scope: INV-KS-2026-05-7102, INV-KS-2026-05-7118.",
                ]),
            ],
            "footer": "Treasury contact: treasury@leewayhertz.com.",
        },
    },
    {
        "slug": "02-hold-release-PO-LH-Q3-3017-credit-limit",
        "sender": "rituraj@leewayhertz.com",
        "subject": "Hold release on PO-LH-Q3-3017 - credit-limit review complete, please release",
        "body": (
            "Hi Keysight Sales Operations,\n\n"
            "Following the credit-limit review on our Leeway Hertz account, our CFO "
            "has confirmed the temporary credit-limit increase to USD 750,000, "
            "effective immediately (approval ref CFO-APPR-2026-06-3017). This brings "
            "PO-LH-Q3-3017 (M9421A VXT Vector Transceiver bundle plus 85052D cal kit, "
            "ship-to Hyderabad RF Lab) back within available credit.\n\n"
            "Please lift the credit hold on PO-LH-Q3-3017 so the order can ship to the "
            "current EndDate on the order. The classified comms validation suite the "
            "VXT supports is on a federal-contract milestone, so the lab cannot absorb "
            "any further delay.\n\n"
            "Attached are the CFO approval memo and the credit-limit update summary "
            "from our treasury team.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "CreditLimit_update_PO-LH-Q3-3017.xlsx",
            "sheet_name": "Credit limit update",
            "title": "Leeway Hertz - Credit limit update for PO-LH-Q3-3017",
            "header": ["Date", "Action", "Previous Limit (USD)", "New Limit (USD)", "Approver", "Reference"],
            "rows": [
                ["2026-06-20", "Initial credit-limit review opened", "500,000", "500,000", "Treasury", "CR-OPEN-LH-2026-06"],
                ["2026-06-26", "CFO interim review",                "500,000", "500,000", "Sanjay Iyer (CFO)", "CR-REVIEW-LH-2026-06"],
                ["2026-06-28", "Credit-limit increase approved",     "500,000", "750,000", "Sanjay Iyer (CFO)", "CFO-APPR-2026-06-3017"],
                ["2026-06-29", "Available headroom after PO-LH-Q3-3017", "500,000", "612,400", "Treasury", "CR-CONFIRM-LH-2026-06"],
            ],
            "footer": "Credit limit raised to USD 750,000 effective 2026-06-28. PO-LH-Q3-3017 now within available credit.",
        },
        "pdf": {
            "filename": "CFO_credit_approval_PO-LH-Q3-3017.pdf",
            "title": "Leeway Hertz - CFO credit-limit approval memo",
            "intent": "Internal CFO memo authorising temporary credit-limit increase covering PO-LH-Q3-3017.",
            "sections": [
                ("Approval", [
                    "Reference: CFO-APPR-2026-06-3017",
                    "Approver: Sanjay Iyer, CFO, Leeway Hertz",
                    "Effective date: 2026-06-28",
                    "Action: Credit limit with Keysight raised from USD 500,000 to USD 750,000.",
                ]),
                ("Rationale", [
                    "PO-LH-Q3-3017 (M9421A VXT bundle) supports a federal-contract milestone in the Hyderabad RF Lab.",
                    "Order value plus existing open balance briefly exceeded the prior credit limit during Q3 capex closeout.",
                    "Treasury confirmed receivables coverage and recommended a temporary increase to absorb the seasonal spike.",
                ]),
                ("Scope", [
                    "Increase is temporary, valid through 2026-12-31.",
                    "Standard 60-day-net payment terms remain in force.",
                    "No change to dispute resolution or escalation clauses in MA-LH-KS-2025-04-01.",
                ]),
            ],
            "footer": "Authorised under section 7.4 of the Leeway Hertz Treasury Policy.",
        },
    },
]


def write_email_txt(folder: Path, demo: dict) -> Path:
    path = folder / "email.txt"
    body = (
        f"FROM:    {demo['sender']}\n"
        f"TO:      [Keysight sandbox mailbox the demo instance is polling]\n"
        f"SUBJECT: {demo['subject']}\n\n"
        f"{demo['body']}\n"
    )
    path.write_text(body, encoding="utf-8")
    return path


def write_excel(folder: Path, demo: dict) -> Path:
    spec = demo["excel"]
    path = folder / spec["filename"]
    wb = Workbook()
    ws = wb.active
    ws.title = spec["sheet_name"][:31]
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Side(border_style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=1, value=spec["title"]).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(spec["header"]))

    header_row = 3
    for col_idx, h in enumerate(spec["header"], start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    for r_idx, row in enumerate(spec["rows"], start=header_row + 1):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = border
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")

    footer_row = header_row + 1 + len(spec["rows"]) + 1
    ws.cell(row=footer_row, column=1, value=spec.get("footer", "")).font = Font(italic=True, color="6B7280")
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(spec["header"]))

    for col_idx in range(1, len(spec["header"]) + 1):
        letter = ws.cell(row=header_row, column=col_idx).column_letter
        max_len = len(str(spec["header"][col_idx - 1]))
        for row in spec["rows"]:
            v = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 38)

    wb.save(path)
    return path


def write_pdf(folder: Path, demo: dict) -> Path:
    spec = demo["pdf"]
    path = folder / spec["filename"]
    doc = SimpleDocTemplate(
        str(path),
        pagesize=LETTER,
        leftMargin=0.85 * inch,
        rightMargin=0.85 * inch,
        topMargin=0.85 * inch,
        bottomMargin=0.85 * inch,
        title=spec["title"],
        author=demo.get("sender", "Procurement"),
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14, leading=18, spaceAfter=10, textColor=colors.HexColor("#1F4E79"))
    intent = ParagraphStyle("Intent", parent=styles["Italic"], fontSize=9.5, leading=13, textColor=colors.HexColor("#6B7280"), spaceAfter=14)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=11.5, leading=16, spaceBefore=10, spaceAfter=4, textColor=colors.HexColor("#111827"))
    body = ParagraphStyle("Body", parent=styles["BodyText"], fontSize=10.5, leading=15, spaceAfter=3, textColor=colors.HexColor("#111827"))
    footer = ParagraphStyle("Footer", parent=styles["Italic"], fontSize=9, leading=12, textColor=colors.HexColor("#6B7280"), spaceBefore=18)

    story = []
    story.append(Paragraph(spec["title"], h1))
    if spec.get("intent"):
        story.append(Paragraph(spec["intent"], intent))
    for section_title, bullets in spec["sections"]:
        story.append(Paragraph(section_title, h2))
        for b in bullets:
            story.append(Paragraph(f"&bull; {b}", body))
    if spec.get("footer"):
        story.append(Paragraph(spec["footer"], footer))
    doc.build(story)
    return path


def write_readme(folder: Path, demo: dict) -> Path:
    path = folder / "README.txt"
    body = (
        f"Demo pack: {demo['slug']}\n"
        f"Expected intent: hold_release\n"
        f"Sender: {demo['sender']}\n"
        f"{'=' * 60}\n\n"
        f"Salesforce backing:\n"
        f"  The referenced PO is already in Salesforce as a Leeway Hertz Order\n"
        f"  tagged with OrderReferenceNumber LIKE 'HOLD-%'. Stage 2.4 fetches\n"
        f"  this as `orders_on_hold` and Stage 5 grounds the reply on it.\n\n"
        f"How to run this case:\n"
        f"  1. From your email client signed in as the sender above,\n"
        f"     compose a new message to the demo's polling mailbox.\n"
        f"  2. Paste the subject and body from email.txt.\n"
        f"  3. Attach BOTH files in this folder.\n"
        f"  4. Send the email; IMAP picks it up within ten seconds.\n\n"
        f"Expected flow: Stage 1 classifies as hold_release. Stage 2 reads the\n"
        f"payment proof and resolves the on-hold Order in Salesforce. Stage 4\n"
        f"lifts the credit hold; Stage 5 drafts the confirmation reply with\n"
        f"the SF Order number, EndDate, and hold reason quoted from SF.\n\n"
        f"Subject (copy verbatim):\n  {demo['subject']}\n"
    )
    path.write_text(body, encoding="utf-8")
    return path


def main() -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    print(f"Writing demo packs under {OUTPUT_ROOT}")
    for demo in DEMOS:
        folder = OUTPUT_ROOT / demo["slug"]
        folder.mkdir(parents=True, exist_ok=True)
        write_email_txt(folder, demo)
        write_excel(folder, demo)
        write_pdf(folder, demo)
        write_readme(folder, demo)
        print(f"  [{demo['slug']}] email.txt + {demo['excel']['filename']} + {demo['pdf']['filename']}")

    idx = OUTPUT_ROOT / "INDEX.txt"
    lines = [
        "Leeway Hertz hold-release demo packs",
        "=" * 60,
        "",
        "Two demo cases targeting orders that are tagged on-hold in",
        "Salesforce (OrderReferenceNumber LIKE 'HOLD-%').",
        "",
        "Salesforce sandbox prerequisites (already provisioned):",
        "  Account: Leeway Hertz (LEEWAY-HERTZ-001 / 001dM00003moontQAA)",
        "  Contact: rituraj@leewayhertz.com (003dM00001xu0FlQAI)",
        "  Orders on hold:",
        "    OrderNumber 00000138, PO PO-LH-Q3-3014, HOLD-OVERDUE-INVOICE,",
        "        N9030B PXA 50 GHz + Option B85, Pune RF Lab, EndDate 2026-07-18",
        "    OrderNumber 00000139, PO PO-LH-Q3-3017, HOLD-CREDIT-LIMIT,",
        "        M9421A VXT bundle + 85052D, Hyderabad RF Lab, EndDate 2026-07-29",
        "",
        "Demo cases:",
    ]
    for d in DEMOS:
        lines.append(f"  - {d['slug']:50s} {d['subject']}")
    lines.append("")
    lines.append("Each subfolder contains email.txt + Excel + PDF + README.txt.")
    idx.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {idx}")


if __name__ == "__main__":
    main()
