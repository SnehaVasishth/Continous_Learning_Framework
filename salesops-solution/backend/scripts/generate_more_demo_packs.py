"""Generate five more demo packs under ``~/Downloads/leewayhertz-more-demo-packs/``:

  01  Leeway Hertz, general_inquiry — status check on PO-LH-Q3-1817
  02  Leeway Hertz, hold_release    — payment cleared for PO-LH-Q2-1714
  03  Leeway Hertz, delivery_change — push PO-LH-Q3-2121 ship to October
  04  Leeway Hertz, general_inquiry — ETA confirmation for PO-LH-Q3-1823 (the slip)
  05  Bluehawk Defense, KSO routing — federal-prime customer (redirected to keysightorders@)

Each pack maps to data already seeded in Salesforce so the reply drafter has
authoritative numbers to quote.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate


OUTPUT_ROOT = Path.home() / "Downloads" / "leewayhertz-more-demo-packs"


DEMOS = [
    {
        "slug": "01-LH-status-check-PO-LH-Q3-1817",
        "intent_hint": "general_inquiry",
        "sender": "rituraj@leewayhertz.com",
        "subject": "Status check on PO-LH-Q3-1817 — confirm current Keysight EndDate",
        "body": (
            "Hi Keysight Trade Order team,\n\n"
            "Could you please confirm the current Keysight-side status and committed "
            "EndDate on our single open order PO-LH-Q3-1817 (N9020A MXA with Option B25, "
            "ship-to Pune Lab B)? Our internal stage gate review is on 18 June and we "
            "need to validate the date that is live in Keysight's system against the "
            "expected delivery on our register.\n\n"
            "Please reply with the Salesforce OrderNumber, current Status, EffectiveDate, "
            "EndDate, and any open issues flagged on the order.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "LH_internal_status_register_PO-LH-Q3-1817.xlsx",
            "sheet_name": "PO status register",
            "title": "Leeway Hertz - Internal status register for PO-LH-Q3-1817",
            "header": ["PO Number", "Line", "Part Number", "Description", "Qty", "Lab", "Internal Expected EndDate"],
            "rows": [
                ["PO-LH-Q3-1817", 1, "N9020A-MXA-3.6", "MXA Signal Analyzer, 3.6 GHz preset", 1, "Pune Lab B", "2026-07-25"],
                ["PO-LH-Q3-1817", 2, "N9020A-OPT-B25", "Wideband Digital Demodulation", 1, "Pune Lab B", "2026-07-25"],
            ],
            "footer": "Stage-gate review 2026-06-18. We need authoritative confirmation against the Keysight-side EndDate by Friday.",
        },
        "pdf": {
            "filename": "LH_stage_gate_review_PO-LH-Q3-1817.pdf",
            "title": "Leeway Hertz - Stage-gate review note for PO-LH-Q3-1817",
            "intent": "Internal Pune Lab B acceptance criteria for the N9020A unit under PO-LH-Q3-1817.",
            "sections": [
                ("Order in scope", [
                    "PO-LH-Q3-1817, Pune Lab B",
                    "N9020A MXA Signal Analyzer with Option B25",
                ]),
                ("Acceptance criteria", [
                    "Calibration certificate dated within 12 months of ship.",
                    "Option B25 factory-installed.",
                    "On dock by 2026-07-25 at the latest to meet the 18 June stage-gate review.",
                ]),
            ],
            "footer": "Contact: rituraj@leewayhertz.com (Procurement Lead).",
        },
    },
    {
        "slug": "02-LH-hold-release-PO-LH-Q2-1714",
        "intent_hint": "hold_release",
        "sender": "rituraj@leewayhertz.com",
        "subject": "Hold release requested on PO-LH-Q2-1714 - wire transfer cleared, please release",
        "body": (
            "Hello Keysight Order Operations,\n\n"
            "The outstanding wire transfer against PO-LH-Q2-1714 (Order in Salesforce for "
            "the E36313B + 34461A bundle, ship-to Pune Power Lab) cleared yesterday. "
            "Our treasury team's wire reference is WIRE-LH-2026-06-2104, value date "
            "2026-06-12. Total settled USD 21,485 against the original invoice.\n\n"
            "Please release the credit hold on this order today so the units can ship "
            "per the EndDate already on the order. The Pune Power Lab is queuing a "
            "calibration acceptance window the day after delivery, so any further hold "
            "will slip our calibration schedule.\n\n"
            "Attached are the wire confirmation and the payment-proof worksheet from "
            "our treasury team.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "Payment_proof_PO-LH-Q2-1714.xlsx",
            "sheet_name": "Wire payment ledger",
            "title": "Leeway Hertz - Wire payment proof for PO-LH-Q2-1714",
            "header": ["Keysight Invoice", "Our PO Reference", "Amount (USD)", "Bank Wire Reference", "Value Date", "Status"],
            "rows": [
                ["INV-KS-2026-05-7102", "PO-LH-Q2-1714", "18,950.00", "WIRE-LH-2026-06-2104", "2026-06-12", "Cleared"],
                ["INV-KS-2026-05-7118", "PO-LH-Q2-1714",  "2,535.00", "WIRE-LH-2026-06-2104", "2026-06-12", "Cleared"],
                ["TOTAL", "PO-LH-Q2-1714", "21,485.00", "WIRE-LH-2026-06-2104", "2026-06-12", "Cleared"],
            ],
            "footer": "PO-LH-Q2-1714 invoices fully settled in this wire batch. Please release hold today.",
        },
        "pdf": {
            "filename": "SWIFT_confirmation_PO-LH-Q2-1714.pdf",
            "title": "SWIFT MT103 confirmation - Leeway Hertz to Keysight",
            "intent": "Outbound wire confirmation against PO-LH-Q2-1714.",
            "sections": [
                ("Wire instruction", [
                    "Sender: Leeway Hertz Pvt Ltd",
                    "Beneficiary: Keysight Technologies Singapore Pte Ltd",
                    "Amount: USD 21,485.00",
                    "Value date: 2026-06-12",
                    "Customer reference: WIRE-LH-2026-06-2104",
                ]),
                ("Application", [
                    "Apply against Order under PO-LH-Q2-1714 and release credit hold.",
                    "Invoices in scope: INV-KS-2026-05-7102 and INV-KS-2026-05-7118.",
                ]),
            ],
            "footer": "Treasury contact: treasury@leewayhertz.com.",
        },
    },
    {
        "slug": "03-LH-delivery-change-PO-LH-Q3-2121",
        "intent_hint": "delivery_change",
        "sender": "rituraj@leewayhertz.com",
        "subject": "Delivery date change on PO-LH-Q3-2121 - push ship to 2026-10-15",
        "body": (
            "Hi Keysight Logistics,\n\n"
            "We need to push the ship date on PO-LH-Q3-2121 (N5230C PNA-L plus 85052D "
            "calibration kit, ship-to Leeway Hertz Hyderabad RF Lab) from the current "
            "committed EndDate to 2026-10-15. Our Hyderabad RF Lab's HVAC retrofit is "
            "running two weeks behind schedule, and the lab will not be ready to receive "
            "the network analyzer until mid-October.\n\n"
            "Please update the order in Salesforce, confirm the new committed ship date, "
            "and acknowledge any cost or warranty-window implications. Bill-to is unchanged "
            "from our standard.\n\n"
            "The attached worksheet has the lab readiness milestones our facilities team "
            "is tracking; the PDF is the formal change request from our procurement office.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "LH_HyderabadRF_lab_readiness.xlsx",
            "sheet_name": "Lab readiness milestones",
            "title": "Leeway Hertz - Hyderabad RF Lab readiness milestones",
            "header": ["Milestone", "Owner", "Original Date", "Revised Date", "Status"],
            "rows": [
                ["HVAC retrofit complete", "Facilities", "2026-08-22", "2026-09-26", "Slipping"],
                ["Power distribution upgrade", "Facilities", "2026-09-05", "2026-10-08", "On track to revised"],
                ["Lab walkthrough + acceptance", "RF Engineering", "2026-09-15", "2026-10-12", "Pending"],
                ["Instrument receive window opens", "Procurement", "2026-09-01", "2026-10-15", "Requested"],
            ],
            "footer": "Net delivery delay required: six weeks. Procurement requesting Keysight reschedule PO-LH-Q3-2121.",
        },
        "pdf": {
            "filename": "LH_formal_change_request_PO-LH-Q3-2121.pdf",
            "title": "Leeway Hertz - Formal change request: PO-LH-Q3-2121 ship date",
            "intent": "Formal procurement request to reschedule PO-LH-Q3-2121 from the current EndDate to 2026-10-15.",
            "sections": [
                ("Change requested", [
                    "PO: PO-LH-Q3-2121",
                    "Product: Keysight N5230C PNA-L plus 85052D calibration kit",
                    "Ship-to: Leeway Hertz Hyderabad RF Lab",
                    "New requested ship date: 2026-10-15",
                ]),
                ("Reason", [
                    "Hyderabad RF Lab HVAC retrofit running two weeks behind schedule.",
                    "Lab will not be ready to receive the network analyzer before 2026-10-12.",
                ]),
                ("Authorisation", [
                    "Authorised by Rituraj Singh, Procurement Lead, Leeway Hertz.",
                    "Reply to confirm the updated EndDate and any cost or warranty implications.",
                ]),
            ],
            "footer": "Reference: LH-CHG-2026-Q3-2121-A. Contact: rituraj@leewayhertz.com.",
        },
    },
    {
        "slug": "04-LH-ETA-confirm-PO-LH-Q3-1823",
        "intent_hint": "general_inquiry",
        "sender": "rituraj@leewayhertz.com",
        "subject": "ETA confirmation on PO-LH-Q3-1823 - Option B25 component allocation status",
        "body": (
            "Hi Keysight Trade Order team,\n\n"
            "We saw on our last status read that PO-LH-Q3-1823 (N9020A MXA with Option B25, "
            "ship-to Pune Lab B) is sitting at an EndDate that is one to two weeks past "
            "the other two N9020A orders in the same batch (PO-LH-Q3-1814 and 1817). Our "
            "understanding is the slip is on the Option B25 component allocation.\n\n"
            "Could you confirm the current Keysight-side EndDate on PO-LH-Q3-1823, the "
            "exact reason for the slip, and whether the date can be pulled in to align "
            "with the other two units? Our Pune Lab B audited acceptance window needs all "
            "three units on dock together.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "LH_three_N9020A_alignment_check.xlsx",
            "sheet_name": "Three-unit alignment",
            "title": "Leeway Hertz - Three-unit alignment for Pune Lab B acceptance",
            "header": ["PO Number", "Unit", "Lab", "Internal Expected EndDate", "Alignment Note"],
            "rows": [
                ["PO-LH-Q3-1814", "N9020A #1", "Pune Lab A", "2026-07-22", "Pacing the batch"],
                ["PO-LH-Q3-1817", "N9020A #2", "Pune Lab B", "2026-07-25", "On the line"],
                ["PO-LH-Q3-1823", "N9020A #3", "Pune Lab B", "2026-07-25", "Reported slip - confirm Keysight side"],
            ],
            "footer": "Audited acceptance window in Pune Lab B requires all three units on dock together.",
        },
        "pdf": {
            "filename": "LH_pune_labB_acceptance_window.pdf",
            "title": "Leeway Hertz - Pune Lab B audited acceptance window",
            "intent": "Internal definition of the audited acceptance window for the Pune Lab B N9020A units.",
            "sections": [
                ("Acceptance window", [
                    "Window opens once all three Pune-bound N9020A units are physically on dock.",
                    "Window is five business days long; metrology team performs paired-instrument acceptance tests.",
                ]),
                ("Dependencies", [
                    "Cannot start with two of three units present; metrology requires all three.",
                    "Any slip on PO-LH-Q3-1823 pushes the entire window for the other two units.",
                ]),
            ],
            "footer": "Owner: Head of Metrology, Leeway Hertz. Contact: metrology@leewayhertz.com.",
        },
    },
    {
        "slug": "05-KSO-bluehawk-defense-status",
        "intent_hint": "kso",
        "sender": "jordan.harlow@bluehawk-defense.com",
        "subject": "Status request - federal-prime PO-BLUEH-DEF-2026-08 and PO-BLUEH-DEF-2026-11",
        "body": (
            "Greetings Keysight Sales Operations,\n\n"
            "Bluehawk Defense Labs Federal Procurement requesting an authoritative status "
            "and committed-delivery read on the two federal-prime orders currently on our "
            "Bluehawk-El Segundo facility:\n\n"
            "  - PO-BLUEH-DEF-2026-08 (N9030B PXA Signal Analyzer 50 GHz with Option B85, "
            "federal radar testbed)\n"
            "  - PO-BLUEH-DEF-2026-11 (M9421A VXT Vector Transceiver bundle, classified "
            "comms validation suite)\n\n"
            "These orders support a federal contract milestone and we are validating the "
            "delivery dates ahead of a government-customer audit. Please confirm the "
            "current Salesforce OrderNumber, Status, and committed EndDate per PO. Route "
            "to keysightorders@keysight.com for federal-prime handling per the existing "
            "arrangement.\n\n"
            "Regards,\n"
            "Jordan Harlow\n"
            "Federal Procurement Officer, Bluehawk Defense Labs\n"
            "jordan.harlow@bluehawk-defense.com"
        ),
        "excel": {
            "filename": "BLUEH_federal_audit_register.xlsx",
            "sheet_name": "Federal audit register",
            "title": "Bluehawk Defense Labs - Federal audit delivery register",
            "header": ["PO Number", "Product", "Federal Contract Milestone", "Required Delivery", "Audit Reviewer"],
            "rows": [
                ["PO-BLUEH-DEF-2026-08", "N9030B PXA 50 GHz + B85", "FED-RADAR-25Q3-M3", "2026-08-22", "DCAA - El Segundo"],
                ["PO-BLUEH-DEF-2026-11", "M9421A VXT bundle", "FED-COMMS-25Q4-M1", "2026-09-12", "DCAA - El Segundo"],
            ],
            "footer": "Federal contract milestones depend on Keysight committed EndDate. Audit walk-in window 2026-08-27 to 2026-08-29.",
        },
        "pdf": {
            "filename": "BLUEH_federal_routing_letter.pdf",
            "title": "Bluehawk Defense Labs - Federal-prime routing letter",
            "intent": "Standing instruction to route Bluehawk Defense Labs federal-prime orders to keysightorders@keysight.com.",
            "sections": [
                ("Customer category", [
                    "Bluehawk Defense Labs - federal-prime contractor.",
                    "Citizenship and access restrictions apply per the federal contract clause.",
                ]),
                ("Routing arrangement", [
                    "All Keysight order traffic from Bluehawk Defense routes to keysightorders@keysight.com.",
                    "Replies must be authored by US-citizen reviewers only.",
                    "Salesforce CCC handles the case but with restricted visibility per RBAC.",
                ]),
                ("Audit cadence", [
                    "Quarterly federal audit at the El Segundo facility.",
                    "Auditor requires authoritative Salesforce OrderNumber, Status, and EndDate per open PO.",
                ]),
            ],
            "footer": "Contact: jordan.harlow@bluehawk-defense.com (Federal Procurement Officer).",
        },
    },
]


def write_email_txt(folder: Path, demo: dict) -> Path:
    path = folder / "email.txt"
    body = (
        f"FROM:    {demo['sender']}\n"
        f"TO:      [Keysight sandbox mailbox the demo instance is polling]\n"
        f"SUBJECT: {demo['subject']}\n\n"
        f"{demo['body']}\n"
    )
    path.write_text(body, encoding="utf-8")
    return path


def write_excel(folder: Path, demo: dict) -> Path:
    spec = demo["excel"]
    path = folder / spec["filename"]
    wb = Workbook()
    ws = wb.active
    ws.title = spec["sheet_name"][:31]
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Side(border_style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=1, value=spec["title"]).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(spec["header"]))

    header_row = 3
    for col_idx, h in enumerate(spec["header"], start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    for r_idx, row in enumerate(spec["rows"], start=header_row + 1):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = border
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")

    footer_row = header_row + 1 + len(spec["rows"]) + 1
    ws.cell(row=footer_row, column=1, value=spec.get("footer", "")).font = Font(italic=True, color="6B7280")
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(spec["header"]))

    for col_idx in range(1, len(spec["header"]) + 1):
        letter = ws.cell(row=header_row, column=col_idx).column_letter
        max_len = len(str(spec["header"][col_idx - 1]))
        for row in spec["rows"]:
            v = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 38)

    wb.save(path)
    return path


def write_pdf(folder: Path, demo: dict) -> Path:
    spec = demo["pdf"]
    path = folder / spec["filename"]
    doc = SimpleDocTemplate(
        str(path),
        pagesize=LETTER,
        leftMargin=0.85 * inch,
        rightMargin=0.85 * inch,
        topMargin=0.85 * inch,
        bottomMargin=0.85 * inch,
        title=spec["title"],
        author=demo.get("sender", "Procurement"),
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14, leading=18, spaceAfter=10, textColor=colors.HexColor("#1F4E79"))
    intent = ParagraphStyle("Intent", parent=styles["Italic"], fontSize=9.5, leading=13, textColor=colors.HexColor("#6B7280"), spaceAfter=14)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=11.5, leading=16, spaceBefore=10, spaceAfter=4, textColor=colors.HexColor("#111827"))
    body = ParagraphStyle("Body", parent=styles["BodyText"], fontSize=10.5, leading=15, spaceAfter=3, textColor=colors.HexColor("#111827"))
    footer = ParagraphStyle("Footer", parent=styles["Italic"], fontSize=9, leading=12, textColor=colors.HexColor("#6B7280"), spaceBefore=18)

    story = []
    story.append(Paragraph(spec["title"], h1))
    if spec.get("intent"):
        story.append(Paragraph(spec["intent"], intent))
    for section_title, bullets in spec["sections"]:
        story.append(Paragraph(section_title, h2))
        for b in bullets:
            story.append(Paragraph(f"&bull; {b}", body))
    if spec.get("footer"):
        story.append(Paragraph(spec["footer"], footer))
    doc.build(story)
    return path


_INTENT_NOTE = {
    "general_inquiry": (
        "Expected flow: Stage 1 classifies as general_inquiry. Stage 2 enriches against\n"
        "Salesforce and brings the open Order rows for Leeway Hertz onto the context.\n"
        "Stage 5 reply quotes Status, EffectiveDate, and EndDate directly from those rows."
    ),
    "hold_release": (
        "Expected flow: Stage 1 classifies as hold_release. Stage 2 reads the payment\n"
        "proof and resolves the Order in Salesforce. Stage 4 lifts the credit hold;\n"
        "Stage 5 drafts the customer confirmation reply with the Order context."
    ),
    "delivery_change": (
        "Expected flow: Stage 1 classifies as delivery_change. Stage 2 extracts the new\n"
        "requested ship date and resolves the Order in Salesforce. Stage 4 updates the\n"
        "Order EndDate; Stage 5 drafts the customer reply with the new committed date."
    ),
    "kso": (
        "Expected flow: Stage 1 classifies as kso (federal-prime / defense routing).\n"
        "The platform redirects the email to keysightorders@keysight.com per the standing\n"
        "arrangement; no Stage 4 automation runs. The case still surfaces on the inbox\n"
        "as 'redirected' so the operator sees the routing decision."
    ),
}


def write_readme(folder: Path, demo: dict) -> Path:
    path = folder / "README.txt"
    intent_note = _INTENT_NOTE.get(demo["intent_hint"], _INTENT_NOTE["general_inquiry"])
    body = (
        f"Demo pack: {demo['slug']}\n"
        f"Expected intent: {demo['intent_hint']}\n"
        f"Sender: {demo['sender']}\n"
        f"{'=' * 60}\n\n"
        f"How to run this case:\n"
        f"  1. From your email client signed in as the sender above,\n"
        f"     compose a new message to the demo's polling mailbox.\n"
        f"  2. Paste the subject and body from email.txt.\n"
        f"  3. Attach BOTH files in this folder.\n"
        f"  4. Send the email; IMAP picks it up within ten seconds.\n\n"
        f"{intent_note}\n\n"
        f"Subject (copy verbatim):\n  {demo['subject']}\n"
    )
    path.write_text(body, encoding="utf-8")
    return path


def main() -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    print(f"Writing demo packs under {OUTPUT_ROOT}")
    for demo in DEMOS:
        folder = OUTPUT_ROOT / demo["slug"]
        folder.mkdir(parents=True, exist_ok=True)
        write_email_txt(folder, demo)
        write_excel(folder, demo)
        write_pdf(folder, demo)
        write_readme(folder, demo)
        print(f"  [{demo['slug']}] {demo['intent_hint']:18s} -> email.txt + {demo['excel']['filename']} + {demo['pdf']['filename']}")

    idx = OUTPUT_ROOT / "INDEX.txt"
    lines = [
        "Leeway Hertz + KSO additional demo packs",
        "=" * 60,
        "",
        "Five demo cases. Each maps to data already seeded in Salesforce so",
        "the Stage 5 reply quotes authoritative Order data, not the customer's",
        "own message.",
        "",
        "Salesforce sandbox prerequisites (already provisioned):",
        "  Leeway Hertz Account (LEEWAY-HERTZ-001) and Contact rituraj@leewayhertz.com",
        "  Bluehawk Defense Labs Account (BLUEH-DEF-021) and Contact jordan.harlow@bluehawk-defense.com",
        "",
        "Demo cases:",
    ]
    for d in DEMOS:
        lines.append(f"  - {d['slug']:42s} [{d['intent_hint']:16s}]  {d['subject']}")
    lines.append("")
    lines.append("Each subfolder contains email.txt + Excel + PDF + README.txt.")
    idx.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {idx}")


if __name__ == "__main__":
    main()
