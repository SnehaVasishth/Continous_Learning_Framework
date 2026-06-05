"""Two work-order-status-inquiry demo cases.

Demo A: dropped in ``~/Downloads/leewayhertz-wo-status-packs/`` so the user
can forward it from their own email client.

Demo B: inserted directly into the local inbox (Email row, attachments
persisted under data/uploads). The IMAP poller surfaces it within ten
seconds and Stage 1 takes it from there.

Both reference WorkOrders already seeded in Salesforce so Stage 2.4's
wo_status_inquiry enrichment pulls real rows for the reply drafter.
"""
from __future__ import annotations

import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate

# Resolve repo + module imports for the inbox-insert path.
HERE = Path(__file__).resolve()
BACKEND = HERE.parent.parent
sys.path.insert(0, str(BACKEND))

from app.db import SessionLocal  # noqa: E402
from app.models import Email     # noqa: E402
from app.config import UPLOADS   # noqa: E402


DOWNLOADS_ROOT = Path.home() / "Downloads" / "leewayhertz-wo-status-packs"


DEMOS = [
    {
        "slug": "01-wo-status-three-WOs-quarterly-review",
        "delivery": "downloads",
        "sender": "rituraj@leewayhertz.com",
        "subject": "Work-order status check - calibration and repair WOs ahead of audit",
        "body": (
            "Hi Keysight Service Operations,\n\n"
            "Ahead of our A2LA audit on Friday, our metrology team needs an authoritative "
            "read on the three open Keysight work orders we have on the Leeway Hertz "
            "account. Please confirm the current Status, expected completion date, "
            "assigned technician, and (where applicable) calibration certificate number "
            "for each of these:\n\n"
            "  - WO-LH-CAL-2026-1417 (N9020A MXA, Pune Lab A, Asset MY52310045)\n"
            "  - WO-LH-REP-2026-2031 (E36313B PSU, Pune Power Lab, Asset MY58370207)\n"
            "  - WO-LH-CAL-2026-2208 (N5230C PNA-L, Hyderabad RF Lab, Asset MY52001108)\n\n"
            "The attached worksheet is our internal status register so you can see the "
            "dates we have on our side; the PDF is the A2LA audit scope letter describing "
            "what the auditor will ask for.\n\n"
            "Please reply with the current Keysight-side values; flag any WO where the "
            "expected completion has slipped from the dates we have on file.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "LH_open_WO_status_register.xlsx",
            "sheet_name": "Open WOs",
            "title": "Leeway Hertz - Open work-order status register",
            "header": ["WO Number", "Asset SKU", "Asset Serial", "Lab", "Type", "Our Expected End", "Status (LH side)"],
            "rows": [
                ["WO-LH-CAL-2026-1417", "N9020A-MXA-3.6", "MY52310045", "Pune Lab A",       "Calibration", "2026-06-03", "Pending Keysight read"],
                ["WO-LH-REP-2026-2031", "E36313B",        "MY58370207", "Pune Power Lab",   "Repair",      "2026-06-12", "Reported held - need cause"],
                ["WO-LH-CAL-2026-2208", "N5230C-PNA-13.5","MY52001108", "Hyderabad RF Lab", "Calibration", "2026-06-24", "Pending Keysight read"],
            ],
            "footer": "Audit window 2026-06-26 to 2026-06-28. All three certificates must be on file before the audit opens.",
        },
        "pdf": {
            "filename": "A2LA_audit_scope_LH_2026.pdf",
            "title": "Leeway Hertz - A2LA audit scope letter (Q2 2026)",
            "intent": "Audit scope letter from our A2LA accreditation body covering the three labs in scope.",
            "sections": [
                ("Audit dates", ["Opening: 2026-06-26", "Closing: 2026-06-28"]),
                ("Labs in scope", [
                    "Leeway Hertz - Pune Lab A (RF spectrum analyzers)",
                    "Leeway Hertz - Pune Power Lab (DC instruments)",
                    "Leeway Hertz - Hyderabad RF Lab (microwave network analyzers)",
                ]),
                ("Auditor evidence required per asset", [
                    "Active calibration certificate within validity window",
                    "Work-order record linking the cert to the asset serial",
                    "Technician name and Keysight calibration laboratory code",
                ]),
            ],
            "footer": "Reference: A2LA-SCOPE-LH-2026-Q2. Lead auditor: M. Whittaker.",
        },
    },
    {
        "slug": "02-wo-status-cal-cert-urgency",
        "delivery": "inbox",
        "sender": "rituraj@leewayhertz.com",
        "subject": "Urgent - calibration certificate status for WO-LH-CAL-2026-1417",
        "body": (
            "Hi Keysight Calibration Operations,\n\n"
            "We need the calibration certificate for WO-LH-CAL-2026-1417 (N9020A MXA, "
            "Asset MY52310045, Pune Lab A) released today. Our acceptance window in "
            "Pune Lab A opens tomorrow morning and the instrument cannot be put on the "
            "bench until the active certificate is on file in our metrology system.\n\n"
            "Please confirm:\n"
            "  - Current Status of WO-LH-CAL-2026-1417 in Keysight Service\n"
            "  - Assigned technician and Keysight calibration lab\n"
            "  - Estimated certificate-issue date\n"
            "  - Direct download link to the certificate (or SharePoint deep-link)\n\n"
            "The attached worksheet is the metrology team's certificate tracker; the PDF "
            "is our Pune Lab A acceptance protocol that defines what the auditor will "
            "look for on Day 1.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "LH_metrology_cert_tracker.xlsx",
            "sheet_name": "Cert tracker",
            "title": "Leeway Hertz - Metrology certificate tracker (this week)",
            "header": ["WO Number", "Asset Serial", "Asset SKU", "Lab", "Cert Status (LH side)", "Required By"],
            "rows": [
                ["WO-LH-CAL-2026-1417", "MY52310045", "N9020A-MXA-3.6", "Pune Lab A", "Awaiting Keysight upload", "2026-06-04 (Lab A acceptance)"],
            ],
            "footer": "Single-WO certificate tracker. Lab A acceptance test cannot start until the cert is on file.",
        },
        "pdf": {
            "filename": "LH_PuneLabA_acceptance_protocol.pdf",
            "title": "Leeway Hertz - Pune Lab A acceptance protocol",
            "intent": "Internal Pune Lab A acceptance protocol for inbound Keysight RF instruments.",
            "sections": [
                ("Day-1 evidence required", [
                    "Active Keysight calibration certificate (PDF) loaded in metrology system.",
                    "Work-order record linking certificate to asset serial.",
                    "Technician + Keysight lab code on the certificate.",
                ]),
                ("Failure mode", [
                    "If cert is not on file by Day 1, instrument is quarantined and the lab acceptance window slips by one calendar week.",
                ]),
            ],
            "footer": "Owner: Head of Metrology, Leeway Hertz. Contact: metrology@leewayhertz.com.",
        },
    },
]


def write_email_txt(folder: Path, demo: dict) -> Path:
    path = folder / "email.txt"
    body = (
        f"FROM:    {demo['sender']}\n"
        f"TO:      [Keysight sandbox mailbox the demo instance is polling]\n"
        f"SUBJECT: {demo['subject']}\n\n"
        f"{demo['body']}\n"
    )
    path.write_text(body, encoding="utf-8")
    return path


def build_xlsx(spec: dict, dest: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = spec["sheet_name"][:31]
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Side(border_style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.cell(row=1, column=1, value=spec["title"]).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(spec["header"]))
    header_row = 3
    for col_idx, h in enumerate(spec["header"], start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center"); c.border = border
    for r_idx, row in enumerate(spec["rows"], start=header_row + 1):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = border
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")
    footer_row = header_row + 1 + len(spec["rows"]) + 1
    ws.cell(row=footer_row, column=1, value=spec.get("footer", "")).font = Font(italic=True, color="6B7280")
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(spec["header"]))
    for col_idx in range(1, len(spec["header"]) + 1):
        letter = ws.cell(row=header_row, column=col_idx).column_letter
        max_len = len(str(spec["header"][col_idx - 1]))
        for row in spec["rows"]:
            v = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 38)
    wb.save(dest)
    return dest


def build_pdf(spec: dict, dest: Path) -> Path:
    doc = SimpleDocTemplate(
        str(dest), pagesize=LETTER,
        leftMargin=0.85 * inch, rightMargin=0.85 * inch,
        topMargin=0.85 * inch, bottomMargin=0.85 * inch,
        title=spec["title"], author="Leeway Hertz Procurement",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14, leading=18, spaceAfter=10, textColor=colors.HexColor("#1F4E79"))
    intent = ParagraphStyle("Intent", parent=styles["Italic"], fontSize=9.5, leading=13, textColor=colors.HexColor("#6B7280"), spaceAfter=14)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=11.5, leading=16, spaceBefore=10, spaceAfter=4, textColor=colors.HexColor("#111827"))
    body = ParagraphStyle("Body", parent=styles["BodyText"], fontSize=10.5, leading=15, spaceAfter=3, textColor=colors.HexColor("#111827"))
    footer = ParagraphStyle("Footer", parent=styles["Italic"], fontSize=9, leading=12, textColor=colors.HexColor("#6B7280"), spaceBefore=18)
    story = [Paragraph(spec["title"], h1)]
    if spec.get("intent"):
        story.append(Paragraph(spec["intent"], intent))
    for section_title, bullets in spec["sections"]:
        story.append(Paragraph(section_title, h2))
        for b in bullets:
            story.append(Paragraph(f"&bull; {b}", body))
    if spec.get("footer"):
        story.append(Paragraph(spec["footer"], footer))
    doc.build(story)
    return dest


def write_readme(folder: Path, demo: dict) -> Path:
    path = folder / "README.txt"
    body = (
        f"Demo pack: {demo['slug']}\n"
        f"Expected intent: wo_status_inquiry\n"
        f"Sender: {demo['sender']}\n"
        f"{'=' * 60}\n\n"
        f"Salesforce backing:\n"
        f"  Three Leeway Hertz WorkOrders are seeded in Salesforce:\n"
        f"    WO-LH-CAL-2026-1417  In Progress  Calibration  N9020A MXA (MY52310045) Pune Lab A\n"
        f"    WO-LH-REP-2026-2031  On Hold      Repair       E36313B PSU (MY58370207) Pune Power Lab\n"
        f"    WO-LH-CAL-2026-2208  New          Calibration  N5230C PNA-L (MY52001108) Hyderabad RF Lab\n"
        f"  Stage 2.4 fetches them as recent_work_orders so Stage 5 grounds the\n"
        f"  reply on real Salesforce data (Status, StartDate, EndDate, Technician).\n\n"
        f"How to run:\n"
        f"  1. From your email client signed in as {demo['sender']}, compose a new\n"
        f"     message to the demo's polling mailbox.\n"
        f"  2. Paste subject and body from email.txt.\n"
        f"  3. Attach BOTH files in this folder.\n"
        f"  4. Send. IMAP picks it up within ten seconds.\n\n"
        f"Subject (copy verbatim):\n  {demo['subject']}\n"
    )
    path.write_text(body, encoding="utf-8")
    return path


def deliver_to_downloads(demo: dict) -> None:
    folder = DOWNLOADS_ROOT / demo["slug"]
    folder.mkdir(parents=True, exist_ok=True)
    write_email_txt(folder, demo)
    build_xlsx(demo["excel"], folder / demo["excel"]["filename"])
    build_pdf(demo["pdf"], folder / demo["pdf"]["filename"])
    write_readme(folder, demo)
    print(f"  [downloads] {demo['slug']} -> {folder}")


def deliver_to_inbox(demo: dict) -> None:
    """Insert the email row + persist attachments under data/uploads, status='new'."""
    UPLOADS.mkdir(parents=True, exist_ok=True)
    slug = demo["slug"]
    xlsx_name = demo["excel"]["filename"]
    pdf_name = demo["pdf"]["filename"]
    xlsx_path = UPLOADS / f"demo_{slug}_{xlsx_name}"
    pdf_path = UPLOADS / f"demo_{slug}_{pdf_name}"
    build_xlsx(demo["excel"], xlsx_path)
    build_pdf(demo["pdf"], pdf_path)
    db = SessionLocal()
    try:
        # The Email.attachments column stores a list of paths the orchestrator
        # later opens. Persist the resolved file paths so the Stage 2 OCR can
        # actually read the XLSX + PDF off disk.
        row = Email(
            received_at=datetime.now(timezone.utc),
            from_address=demo["sender"],
            subject=demo["subject"],
            body=demo["body"],
            language_hint="en",
            attachments=[str(xlsx_path), str(pdf_path)],
            status="new",
            account_id=1,
        )
        db.add(row); db.commit()
        print(f"  [inbox]     {demo['slug']} -> Email id={row.id}, attachments saved under data/uploads/")
    finally:
        db.close()


def main() -> None:
    DOWNLOADS_ROOT.mkdir(parents=True, exist_ok=True)
    for demo in DEMOS:
        if demo["delivery"] == "downloads":
            deliver_to_downloads(demo)
        else:
            deliver_to_inbox(demo)
    # Index file for the Downloads pack
    idx = DOWNLOADS_ROOT / "INDEX.txt"
    lines = [
        "Leeway Hertz - work-order status-inquiry demo packs",
        "=" * 60,
        "",
        "Salesforce backing (already provisioned):",
        "  WO-LH-CAL-2026-1417  In Progress  Calibration  N9020A MXA  Pune Lab A",
        "  WO-LH-REP-2026-2031  On Hold      Repair       E36313B PSU Pune Power Lab",
        "  WO-LH-CAL-2026-2208  New          Calibration  N5230C PNA-L Hyderabad RF Lab",
        "",
        "Downloads pack(s):",
    ]
    for d in DEMOS:
        if d["delivery"] == "downloads":
            lines.append(f"  - {d['slug']:42s} {d['subject']}")
    lines.append("")
    lines.append("A second case (02-wo-status-cal-cert-urgency) has been inserted")
    lines.append("directly into the local inbox (Email status='new') so the IMAP poller")
    lines.append("picks it up within ten seconds and triggers a fresh pipeline.")
    idx.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {idx}")


if __name__ == "__main__":
    main()
