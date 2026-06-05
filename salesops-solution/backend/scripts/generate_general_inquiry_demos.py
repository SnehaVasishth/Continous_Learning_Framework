"""Generate 5 ready-to-send Leeway Hertz demo packs under
``~/Downloads/leewayhertz-demo-packs/``.

Intent mix (deliberate — covers the three flows the demo shows off):
  01  general_inquiry  — lead time question
  02  general_inquiry  — EOL roadmap clarification
  03  hold_release     — payment cleared, customer asks for hold release
  04  po_intake        — fresh PO for new hardware
  05  quote_to_order   — accept Keysight quote and convert to order

Each pack is a self-contained folder:
  - ``email.txt`` — paste-ready subject + body for the user to send from
    rituraj@leewayhertz.com (the SF Contact added to the Leeway Hertz
    account, Contact Id 003dM00001xu0FlQAI).
  - One ``.xlsx`` and one ``.pdf`` attachment, named with intent in mind
    (payment proof for hold release, BOM for PO intake, spec sheet for
    inquiry, etc.).

Run from the backend dir:
    .venv/bin/python scripts/generate_general_inquiry_demos.py
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib import colors


OUTPUT_ROOT = Path.home() / "Downloads" / "leewayhertz-demo-packs"


# ---------------------------------------------------------------------------
# Demo definitions — 5 Salesforce-aware general-inquiry scenarios from a
# Leeway Hertz buyer. Each pack carries one Excel + one PDF attachment.
# ---------------------------------------------------------------------------

DEMOS: list[dict] = [
    {
        "slug": "01-lead-time-n9020a",
        "subject": "Lead time confirmation needed — three N9020A MXA units for Q3 production",
        "body": (
            "Hello Keysight Trade Order team,\n\n"
            "We are finalising our Q3 production schedule and need a firm lead time read on three Keysight N9020A MXA "
            "Signal Analyzers (3.6 GHz preset). Our internal stage gate closes 18 June, so we need confirmation that "
            "all three units can be on our dock at our Pune lab by 25 July at the latest.\n\n"
            "The attached worksheet lists the three open POs we are tracking on our side, with the line numbers and "
            "the requested delivery dates. The PDF is the N9020A configuration we are standardising on across the lab "
            "(the Option B25 wideband demod is mandatory, the rest are nice-to-have).\n\n"
            "Could you confirm two things by Friday: (1) the manufacturing lead time for the three units, including "
            "Option B25 build, and (2) whether all three can ship together from the Singapore hub or whether one will "
            "split across hubs. We are happy to consolidate freight if it shortens the window.\n\n"
            "Best regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "Open_PO_lead_time_tracker.xlsx",
            "sheet_name": "Q3-2026 N9020A POs",
            "title": "Leeway Hertz — Open PO lead time tracker (Q3 2026)",
            "header": ["PO Number", "Line", "Part Number", "Description", "Qty", "Requested Delivery", "Ship-To", "Notes"],
            "rows": [
                ["PO-LH-Q3-1814", 1, "N9020A-MXA-3.6", "MXA Signal Analyzer, 3.6 GHz preset", 1, "2026-07-21", "Pune Lab A", "B25 required"],
                ["PO-LH-Q3-1814", 2, "N9020A-OPT-B25", "Wideband Digital Demodulation", 1, "2026-07-21", "Pune Lab A", "Build-to-order"],
                ["PO-LH-Q3-1817", 1, "N9020A-MXA-3.6", "MXA Signal Analyzer, 3.6 GHz preset", 1, "2026-07-25", "Pune Lab B", "B25 required"],
                ["PO-LH-Q3-1817", 2, "N9020A-OPT-B25", "Wideband Digital Demodulation", 1, "2026-07-25", "Pune Lab B", "Ship with line 1"],
                ["PO-LH-Q3-1823", 1, "N9020A-MXA-3.6", "MXA Signal Analyzer, 3.6 GHz preset", 1, "2026-07-25", "Pune Lab B", "B25 required"],
                ["PO-LH-Q3-1823", 2, "N9020A-OPT-B25", "Wideband Digital Demodulation", 1, "2026-07-25", "Pune Lab B", "Consolidate freight OK"],
            ],
            "footer": (
                "Stage gate closes 2026-06-18. All three units must be on dock at Pune by 2026-07-25. "
                "Consolidated freight from Singapore hub preferred over split shipments."
            ),
        },
        "pdf": {
            "filename": "N9020A_standard_config_LeewayHertz.pdf",
            "title": "Leeway Hertz — Standard N9020A MXA Configuration (Lab rollout, Q3 2026)",
            "intent": "Internal procurement standard for all new N9020A units acquired in 2026.",
            "sections": [
                ("Base instrument", [
                    "Model: Keysight N9020A MXA Signal Analyzer",
                    "Frequency range: 20 Hz to 3.6 GHz (preset)",
                    "Resolution bandwidth: 1 Hz to 8 MHz",
                ]),
                ("Mandatory options", [
                    "Option B25 — Wideband Digital Demodulation (mandatory across all new units)",
                    "Option EXM — External mixing capability",
                    "3-year calibration contract (CalSure Gold)",
                ]),
                ("Nice-to-have", [
                    "Option H1G — 1 GHz analysis bandwidth (case-by-case)",
                    "Option N9063A — Analog modulation analysis software",
                ]),
                ("Sites in scope", [
                    "Pune Lab A — N9020A x 1 (Q3 2026)",
                    "Pune Lab B — N9020A x 2 (Q3 2026)",
                    "Hyderabad RF Lab — N9020A x 1 (Q4 2026)",
                ]),
                ("Acceptance gate", [
                    "Unit must arrive with current Keysight calibration certificate, valid 12 months from ship date.",
                    "Option B25 must be factory-installed; field-install is not accepted for this rollout.",
                ]),
            ],
            "footer": (
                "Document owner: Procurement Lead, Leeway Hertz. Last revised 2026-05-15. "
                "Contact: rituraj@leewayhertz.com."
            ),
        },
    },
    {
        "slug": "02-eol-roadmap-e36312a",
        "subject": "EOL roadmap clarification — Keysight E36312A power supplies in our fleet",
        "body": (
            "Hi Keysight Support,\n\n"
            "Our calibration team flagged the recent EOL notice for the E36312A triple-output power supply during our "
            "last asset review. We have eleven E36312A units in active service across two labs, with the oldest five "
            "approaching their next cal cycle in September.\n\n"
            "Before we open replacement POs we need clarity on the roadmap. The attached asset register lists every "
            "E36312A we currently own, serial, lab assignment, last cal date, and next cal due. The PDF is the EOL "
            "announcement we received internally — we want to make sure we are reading the spare-parts support window "
            "correctly.\n\n"
            "Three questions: (1) Is the published 31-Dec-2027 last-time-buy date final, or is there a chance it gets "
            "extended for enterprise accounts? (2) What is the spare-parts and calibration support window beyond "
            "last-time-buy? And (3) is the E36300 series the recommended migration target, or are you steering "
            "enterprise customers somewhere else?\n\n"
            "Thanks,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "E36312A_asset_register.xlsx",
            "sheet_name": "E36312A units",
            "title": "Leeway Hertz — E36312A power supply asset register",
            "header": ["Asset Tag", "Serial Number", "Lab", "Acquired", "Last Cal", "Next Cal Due", "Cal Provider", "Status"],
            "rows": [
                ["LH-PS-0142", "MY58370112", "Pune RF Lab",    "2022-03-14", "2025-09-08", "2026-09-08", "Keysight CalSure", "Active"],
                ["LH-PS-0143", "MY58370118", "Pune RF Lab",    "2022-03-14", "2025-09-12", "2026-09-12", "Keysight CalSure", "Active"],
                ["LH-PS-0151", "MY58370207", "Pune Power Lab", "2022-08-02", "2025-10-21", "2026-10-21", "Keysight CalSure", "Active"],
                ["LH-PS-0152", "MY58370214", "Pune Power Lab", "2022-08-02", "2025-10-21", "2026-10-21", "Keysight CalSure", "Active"],
                ["LH-PS-0166", "MY58370341", "Hyderabad RF",   "2023-01-19", "2026-01-22", "2027-01-22", "Keysight CalSure", "Active"],
                ["LH-PS-0167", "MY58370347", "Hyderabad RF",   "2023-01-19", "2026-01-22", "2027-01-22", "Keysight CalSure", "Active"],
                ["LH-PS-0181", "MY58370402", "Hyderabad RF",   "2023-06-08", "2026-06-11", "2027-06-11", "Keysight CalSure", "Active"],
                ["LH-PS-0182", "MY58370407", "Hyderabad RF",   "2023-06-08", "2026-06-11", "2027-06-11", "Keysight CalSure", "Active"],
                ["LH-PS-0204", "MY58370515", "Pune Power Lab", "2024-02-21", "2027-02-22", "2028-02-22", "Keysight CalSure", "Active"],
                ["LH-PS-0205", "MY58370521", "Pune Power Lab", "2024-02-21", "2027-02-22", "2028-02-22", "Keysight CalSure", "Active"],
                ["LH-PS-0206", "MY58370529", "Pune Power Lab", "2024-02-21", "2027-02-22", "2028-02-22", "Keysight CalSure", "Active"],
            ],
            "footer": (
                "Total units in service: 11. Five units approach next cal cycle in September and October 2026 — "
                "decision on replacement vs. continued service needed before 2026-08-15."
            ),
        },
        "pdf": {
            "filename": "Keysight_E36312A_EOL_announcement.pdf",
            "title": "Keysight E36312A Triple-Output Power Supply — End-of-Life Announcement",
            "intent": "Customer-facing EOL notice received via Keysight enterprise account portal on 2026-04-22.",
            "sections": [
                ("Product affected", [
                    "Model: E36312A — Triple-output programmable DC power supply",
                    "Configurations covered: All factory configurations of E36312A",
                    "Successor product: E36300 series (E36312B, E36313B)",
                ]),
                ("Key dates", [
                    "Last-time-buy date: 2027-12-31",
                    "Last shipment date: 2028-03-31",
                    "Cal support last date: 2031-12-31 (3 years post last shipment)",
                    "Spare-parts last date: 2031-12-31",
                ]),
                ("Migration guidance", [
                    "Recommended migration: E36313B (drop-in replacement with extended range)",
                    "Trade-in credit available against new E36300 series purchase through end of 2026",
                    "Customer engineering teams available for transition planning (no charge for enterprise accounts)",
                ]),
                ("Support continuity", [
                    "Existing CalSure contracts remain valid through 2031-12-31.",
                    "Firmware patches limited to security and stability fixes after 2027-12-31.",
                    "Application notes and reference designs remain available indefinitely.",
                ]),
            ],
            "footer": (
                "Issued by Keysight Order Management — questions: enterprise.support@keysight.com. "
                "Customer reference: KS-EOL-2026-E36312A. Issued 2026-04-22."
            ),
        },
    },
    {
        "slug": "03-hold-release-SO-LH-8214",
        "intent_hint": "hold_release",
        "subject": "Request to release credit hold on Order SO-LH-8214 — wire transfer cleared",
        "body": (
            "Hello Keysight Order Operations,\n\n"
            "Our finance team confirmed this morning that the outstanding invoices linked to Order SO-LH-8214 have "
            "cleared on the wire we sent on 18 May. Your finance contact at the Singapore office should be able to "
            "see the funds posted as of close of business yesterday.\n\n"
            "We need this order off credit hold as soon as possible. The site engineering team in Pune is waiting "
            "for the two N9020A units on the order to start an audited acceptance test on Monday morning. Holding "
            "the order another day past Friday will slip our internal Q2 closure.\n\n"
            "I have attached two documents. The Excel is the payment-proof worksheet our treasury team uses to "
            "confirm wire references, with the Keysight invoice numbers, your bank reference, our bank reference, "
            "and the value date. The PDF is the SWIFT confirmation from our bank in case your finance team needs "
            "the raw artefact.\n\n"
            "Please confirm the hold is released and trigger the release-from-hold workflow today. The credit limit "
            "on our account should be back to its normal headroom once these invoices flip to paid.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "Payment_proof_SO-LH-8214.xlsx",
            "sheet_name": "Wire payment ledger",
            "title": "Leeway Hertz — Wire payment proof for Order SO-LH-8214",
            "header": ["Keysight Invoice", "Our PO Reference", "Amount (USD)", "Bank Wire Reference", "Value Date", "Status", "Notes"],
            "rows": [
                ["INV-KS-2026-04412", "PO-LH-Q2-1714", "48,225.00", "WIRE-LH-2026-05-1071", "2026-05-18", "Cleared", "Wire posted at value 2026-05-19"],
                ["INV-KS-2026-04428", "PO-LH-Q2-1714", "12,475.00", "WIRE-LH-2026-05-1071", "2026-05-18", "Cleared", "Same wire batch"],
                ["INV-KS-2026-04501", "PO-LH-Q2-1714",  "3,640.00", "WIRE-LH-2026-05-1071", "2026-05-18", "Cleared", "Same wire batch"],
                ["TOTAL", "PO-LH-Q2-1714", "64,340.00", "WIRE-LH-2026-05-1071", "2026-05-18", "Cleared", "Order SO-LH-8214 fully paid"],
            ],
            "footer": (
                "Wire batch WIRE-LH-2026-05-1071 totalling USD 64,340 cleared at value date 2026-05-18. "
                "Three Keysight invoices retired in this batch. Order SO-LH-8214 is now fully paid."
            ),
        },
        "pdf": {
            "filename": "SWIFT_confirmation_LH_to_Keysight.pdf",
            "title": "SWIFT MT103 confirmation — Leeway Hertz to Keysight Technologies",
            "intent": "Outbound wire confirmation from Leeway Hertz primary banking partner.",
            "sections": [
                ("Wire instruction", [
                    "Sender: Leeway Hertz Pvt Ltd — Account ending 4719",
                    "Beneficiary: Keysight Technologies Singapore Pte Ltd",
                    "Beneficiary bank: Standard Chartered, Singapore (SCBLSGSG)",
                    "Amount: USD 64,340.00",
                    "Value date: 2026-05-18",
                ]),
                ("Reference fields", [
                    "Customer reference: WIRE-LH-2026-05-1071",
                    "Invoice references: INV-KS-2026-04412, INV-KS-2026-04428, INV-KS-2026-04501",
                    "Order: SO-LH-8214",
                    "Purpose: Settlement of Q2 invoices for Keysight Order SO-LH-8214",
                ]),
                ("Execution", [
                    "Initiated: 2026-05-18 09:42 IST",
                    "Status: Settled at value date 2026-05-18",
                    "Acknowledgement: Beneficiary bank ack received 2026-05-19 04:11 SGT",
                ]),
                ("Notes", [
                    "Please apply the entire amount to Order SO-LH-8214 and release credit hold.",
                    "All three invoices in scope are settled in full by this wire batch.",
                ]),
            ],
            "footer": (
                "Issued by Leeway Hertz treasury operations — for questions contact treasury@leewayhertz.com. "
                "This is the official SWIFT confirmation and supersedes any prior payment correspondence."
            ),
        },
    },
    {
        "slug": "04-po-intake-LH-2026-7700",
        "intent_hint": "po_intake",
        "subject": "Purchase Order PO-LH-2026-7700 — N9020B MXA Signal Analyzer x 2 plus N7045A SDR",
        "body": (
            "Hello Keysight Trade Order team,\n\n"
            "Please find attached PO-LH-2026-7700 for our Q3 2026 lab refresh in Pune. The PO covers two N9020B MXA "
            "signal analyzers (with Option H1G one-GHz analysis bandwidth) and one N7045A wideband SDR for our 5G "
            "test rig. Total order value is USD 248,400, payment terms Net 30 against our master agreement.\n\n"
            "The Excel attachment is the line-item BOM with part numbers, configured options, unit price, and target "
            "delivery dates. The PDF is the signed purchase order issued by our procurement department, with the "
            "authorised signatory and the bill-to/ship-to blocks.\n\n"
            "Two notes for the order acceptance pass. First, confirm Option H1G is build-to-order with a 6-week "
            "lead-time so we set internal expectations accordingly. Second, all three units must arrive together on "
            "the same dock at Pune Lab A; please consolidate from the Singapore hub even if the analyzers are ready "
            "earlier than the SDR.\n\n"
            "Acknowledge receipt and confirm the expected ship date by reply. Our standard acceptance window is 5 "
            "business days from your acknowledgement.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "BOM_PO-LH-2026-7700.xlsx",
            "sheet_name": "PO-LH-2026-7700 BOM",
            "title": "Leeway Hertz — Purchase Order BOM (PO-LH-2026-7700)",
            "header": ["Line", "Part Number", "Description", "Qty", "Unit Price (USD)", "Extended (USD)", "Requested Delivery", "Notes"],
            "rows": [
                [1, "N9020B-MXA",     "MXA Signal Analyzer, B-series", 2, "92,400.00", "184,800.00", "2026-08-22", "Build-to-order, Option H1G"],
                [2, "N9020B-OPT-H1G", "1 GHz analysis bandwidth",      2,  "12,800.00",  "25,600.00", "2026-08-22", "Factory-installed only"],
                [3, "N7045A",          "Wideband SDR, 6 GHz",            1,  "38,000.00",  "38,000.00", "2026-08-22", "Standard config"],
                ["", "", "Subtotal", "", "", "248,400.00", "", ""],
                ["", "", "Freight (Singapore consolidated)", "", "", "Included", "", "Tier 2 per master agreement"],
                ["", "", "TOTAL", "", "", "248,400.00", "", "Net 30"],
            ],
            "footer": (
                "PO total USD 248,400 against master agreement MA-LH-KS-2025-04-01. Net 30 payment terms. "
                "All three units to consolidate at Singapore hub and ship together to Pune Lab A by 2026-08-22."
            ),
        },
        "pdf": {
            "filename": "LH_PurchaseOrder_PO-LH-2026-7700_signed.pdf",
            "title": "Purchase Order PO-LH-2026-7700",
            "intent": "Issued by Leeway Hertz Procurement on 2026-05-19. Authorised against master agreement MA-LH-KS-2025-04-01.",
            "sections": [
                ("Buyer", [
                    "Leeway Hertz Pvt Ltd",
                    "Bill-to: Plot 17, Hinjewadi IT Park Phase 2, Pune 411057, India",
                    "Ship-to: Leeway Hertz Pune Lab A, Plot 17, Hinjewadi IT Park Phase 2, Pune 411057",
                    "GSTIN: 27AABCL5432F1Z1",
                    "Account number: LEEWAY-HERTZ-001",
                ]),
                ("Seller", [
                    "Keysight Technologies Singapore Pte Ltd",
                    "Master agreement: MA-LH-KS-2025-04-01",
                    "Payment terms: Net 30 against confirmed delivery",
                ]),
                ("Line summary", [
                    "Line 1: Keysight N9020B MXA Signal Analyzer, B-series — Qty 2 — USD 92,400 each",
                    "Line 2: Option H1G 1 GHz analysis bandwidth (factory-installed) — Qty 2 — USD 12,800 each",
                    "Line 3: Keysight N7045A Wideband SDR 6 GHz — Qty 1 — USD 38,000 each",
                    "Order total: USD 248,400",
                ]),
                ("Acceptance conditions", [
                    "All three units to ship consolidated from the Singapore hub on or before 2026-08-22.",
                    "Each unit to arrive with current Keysight calibration certificate, valid 12 months from ship date.",
                    "Option H1G to be factory-installed; field-install is not accepted on this PO.",
                ]),
                ("Authorised signatory", [
                    "Rituraj Singh — Procurement Lead, Leeway Hertz",
                    "rituraj@leewayhertz.com",
                    "Signed and issued: 2026-05-19",
                ]),
            ],
            "footer": (
                "This purchase order is governed by Master Agreement MA-LH-KS-2025-04-01 and the standard Keysight "
                "terms incorporated therein. Questions to rituraj@leewayhertz.com."
            ),
        },
    },
    {
        "slug": "05-quote-to-order-QT-LH-9920",
        "intent_hint": "quote_to_order",
        "subject": "Accept Keysight quote QT-LH-9920 — convert to firm order, ship to Hyderabad RF",
        "body": (
            "Hi Keysight Sales Operations,\n\n"
            "We are accepting Keysight quote QT-LH-9920 as issued, with no revisions. Please convert it to a firm "
            "order and route to your trade order team for acceptance.\n\n"
            "The Excel attachment is our internal approval workflow record, showing the head-of-engineering and "
            "the CFO sign-off against the quote subtotal and the asset budget line it draws from. The PDF is the "
            "original Keysight quote document with the line items and the quoted unit prices we are accepting.\n\n"
            "Ship-to is our Hyderabad RF Lab. Bill-to is unchanged from our standard. Payment terms Net 30 against "
            "our master agreement. We need acknowledgement of the converted order and the new SO number by reply, "
            "so our finance can pre-load the invoice expectation.\n\n"
            "One ask on timing: the quote validity runs through 2026-06-15 — please target acceptance and order "
            "creation by 2026-06-12 so we have a four-day buffer.\n\n"
            "Regards,\n"
            "Rituraj Singh\n"
            "Procurement Lead, Leeway Hertz\n"
            "rituraj@leewayhertz.com"
        ),
        "excel": {
            "filename": "LH_Internal_Approval_QT-LH-9920.xlsx",
            "sheet_name": "Internal approval",
            "title": "Leeway Hertz — Internal approval workflow for Keysight quote QT-LH-9920",
            "header": ["Step", "Approver", "Role", "Approval Date", "Budget Line", "Amount (USD)", "Status", "Notes"],
            "rows": [
                [1, "Dr. Anika Mehta",     "Head of Engineering",  "2026-05-12", "ENG-CAPEX-2026-Q2", "118,250.00", "Approved", "Aligned with RF lab refresh plan"],
                [2, "Vikram Kapoor",       "Head of Procurement",  "2026-05-14", "ENG-CAPEX-2026-Q2", "118,250.00", "Approved", "Vendor risk check complete"],
                [3, "Sanjay Iyer",         "CFO",                  "2026-05-16", "ENG-CAPEX-2026-Q2", "118,250.00", "Approved", "Within Q2 CAPEX envelope"],
                [4, "Rituraj Singh",        "Procurement Lead (PO issuer)", "2026-05-19", "ENG-CAPEX-2026-Q2", "118,250.00", "Issuing PO", "Acceptance email to Keysight today"],
            ],
            "footer": (
                "Full sign-off chain complete on 2026-05-16. Quote QT-LH-9920 cleared internal review against the "
                "ENG-CAPEX-2026-Q2 envelope. Procurement issuing acceptance to Keysight on 2026-05-19."
            ),
        },
        "pdf": {
            "filename": "Keysight_Quote_QT-LH-9920.pdf",
            "title": "Keysight Sales Quote QT-LH-9920",
            "intent": "Quote issued by Keysight Sales Singapore on 2026-04-28. Valid through 2026-06-15.",
            "sections": [
                ("Quote header", [
                    "Quote number: QT-LH-9920",
                    "Issued: 2026-04-28",
                    "Valid through: 2026-06-15",
                    "Customer: Leeway Hertz Pvt Ltd (account LEEWAY-HERTZ-001)",
                ]),
                ("Line items", [
                    "Line 1: Keysight N5232B PNA-L Microwave Network Analyzer, 13.5 GHz — Qty 1 — USD 78,400",
                    "Line 2: Keysight 85052D 3.5 mm Economy Calibration Kit — Qty 1 — USD 14,650",
                    "Line 3: Keysight U2001A USB Power Sensor, 9 kHz to 6 GHz — Qty 3 — USD 8,400 each (USD 25,200)",
                    "Quote subtotal: USD 118,250",
                ]),
                ("Terms", [
                    "Payment: Net 30 against confirmed delivery",
                    "Ship-from: Singapore hub",
                    "Ship-to: Leeway Hertz Hyderabad RF Lab (default per master agreement)",
                    "Lead time: 5 to 7 weeks from order acceptance",
                ]),
                ("Acceptance instructions", [
                    "To accept this quote as issued, reply with PO reference and sign-off email from authorised buyer.",
                    "To accept with revisions (price, quantity, ship-to), include the revised line item list inline.",
                    "Quote auto-expires on 2026-06-15 if not accepted.",
                ]),
            ],
            "footer": (
                "Issued by Keysight Sales Operations, Singapore — for questions contact your Keysight account team. "
                "Quote reference: QT-LH-9920 / KS-SO-2026-04-28."
            ),
        },
    },
]


# ---------------------------------------------------------------------------
# Renderers
# ---------------------------------------------------------------------------

def write_email_txt(folder: Path, demo: dict) -> Path:
    path = folder / "email.txt"
    body = (
        f"FROM:    rituraj@leewayhertz.com\n"
        f"TO:      [Keysight Trade Order mailbox — paste here]\n"
        f"SUBJECT: {demo['subject']}\n\n"
        f"{demo['body']}\n"
    )
    path.write_text(body, encoding="utf-8")
    return path


def write_excel(folder: Path, demo: dict) -> Path:
    spec = demo["excel"]
    path = folder / spec["filename"]
    wb = Workbook()
    ws = wb.active
    ws.title = spec["sheet_name"][:31]

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Side(border_style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.cell(row=1, column=1, value=spec["title"]).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(spec["header"]))

    # Header row
    header_row = 3
    for col_idx, h in enumerate(spec["header"], start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    # Data rows
    for r_idx, row in enumerate(spec["rows"], start=header_row + 1):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = border
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")

    # Footer
    footer_row = header_row + 1 + len(spec["rows"]) + 1
    ws.cell(row=footer_row, column=1, value=spec.get("footer", "")).font = Font(italic=True, color="6B7280")
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(spec["header"]))

    # Column widths
    for col_idx in range(1, len(spec["header"]) + 1):
        # Use safer column letter; default sample-driven width.
        letter = ws.cell(row=header_row, column=col_idx).column_letter
        max_len = len(str(spec["header"][col_idx - 1]))
        for row in spec["rows"]:
            v = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 38)

    wb.save(path)
    return path


def write_pdf(folder: Path, demo: dict) -> Path:
    spec = demo["pdf"]
    path = folder / spec["filename"]
    doc = SimpleDocTemplate(
        str(path),
        pagesize=LETTER,
        leftMargin=0.85 * inch,
        rightMargin=0.85 * inch,
        topMargin=0.85 * inch,
        bottomMargin=0.85 * inch,
        title=spec["title"],
        author="Leeway Hertz Procurement",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14, leading=18, spaceAfter=10, textColor=colors.HexColor("#1F4E79"))
    intent = ParagraphStyle("Intent", parent=styles["Italic"], fontSize=9.5, leading=13, textColor=colors.HexColor("#6B7280"), spaceAfter=14)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=11.5, leading=16, spaceBefore=10, spaceAfter=4, textColor=colors.HexColor("#111827"))
    body = ParagraphStyle("Body", parent=styles["BodyText"], fontSize=10.5, leading=15, spaceAfter=3, textColor=colors.HexColor("#111827"))
    footer = ParagraphStyle("Footer", parent=styles["Italic"], fontSize=9, leading=12, textColor=colors.HexColor("#6B7280"), spaceBefore=18)

    story = []
    story.append(Paragraph(spec["title"], h1))
    if spec.get("intent"):
        story.append(Paragraph(spec["intent"], intent))
    for section_title, bullets in spec["sections"]:
        story.append(Paragraph(section_title, h2))
        for b in bullets:
            story.append(Paragraph(f"• {b}", body))
    if spec.get("footer"):
        story.append(Paragraph(spec["footer"], footer))
    doc.build(story)
    return path


_INTENT_BLURB = {
    "general_inquiry": (
        "  - Stage 1 intake classifies the email as general_inquiry.\n"
        "  - Stage 2 OCRs the PDF and parses the Excel attachment.\n"
        "  - Stage 2.3 matches Rituraj to the Leeway Hertz Salesforce\n"
        "    account via the Contact record.\n"
        "  - Stage 3 routes the request through the Inquiry track,\n"
        "    no AIOA call (general_inquiry is out of AIOA scope).\n"
        "  - Stage 4 takes the draft_reply action (no SF write).\n"
        "  - Stage 5 drafts the customer-facing answer using the email\n"
        "    content, the SF account context, and the attachments."
    ),
    "hold_release": (
        "  - Stage 1 intake classifies the email as hold_release.\n"
        "  - Stage 2 OCRs the SWIFT confirmation PDF and parses the\n"
        "    payment-proof Excel; the extractor pulls the order number\n"
        "    SO-LH-8214 and the wire reference.\n"
        "  - Stage 2.3 matches Rituraj to the Leeway Hertz account.\n"
        "  - Stage 3 looks up the existing SF Case for this order; if\n"
        "    one is found the LLM duplicate matcher short-circuits the\n"
        "    pipeline. If none is found, a new CCC Request is opened.\n"
        "  - Stage 4 fires the release-from-hold workflow against the\n"
        "    Salesforce Order and posts a CaseComment with the payment\n"
        "    proof links.\n"
        "  - Stage 5 drafts the customer confirmation reply (parks at\n"
        "    HITL one-click; demo mode does not actually transmit)."
    ),
    "po_intake": (
        "  - Stage 1 intake classifies the email as po_intake.\n"
        "  - Stage 2 OCRs the signed-PO PDF and parses the BOM Excel,\n"
        "    extracting line items, unit prices, options, and ship-to.\n"
        "  - Stage 2.3 matches Rituraj to the Leeway Hertz account.\n"
        "  - Stage 3 hands the PO off to AIOA for order acceptance.\n"
        "    When AIOA returns PASS, Stage 4 creates the Salesforce\n"
        "    Order and Order Lines and posts the SOA. When AIOA returns\n"
        "    FAIL the pipeline parks for HITL one-click clarification.\n"
        "  - Stage 5 drafts the order acknowledgement reply with the\n"
        "    SOA attachment and SharePoint link."
    ),
    "quote_to_order": (
        "  - Stage 1 intake classifies the email as quote_to_order.\n"
        "  - Stage 2 OCRs the Keysight quote PDF and parses the internal\n"
        "    approval Excel; the extractor pulls quote number QT-LH-9920\n"
        "    and the accepted line items.\n"
        "  - Stage 2.3 matches Rituraj to the Leeway Hertz account.\n"
        "  - Stage 3 hands the quote acceptance to AIOA for validation.\n"
        "    PASS routes to Stage 4 to convert the quote to a firm Order\n"
        "    in Salesforce. FAIL parks for HITL one-click clarification.\n"
        "  - Stage 5 drafts the customer confirmation reply with the\n"
        "    new SO number once the order is created."
    ),
}


def write_readme(folder: Path, demo: dict) -> Path:
    path = folder / "README.txt"
    excel_name = demo["excel"]["filename"]
    pdf_name = demo["pdf"]["filename"]
    intent_hint = demo.get("intent_hint", "general_inquiry")
    intent_blurb = _INTENT_BLURB.get(intent_hint, _INTENT_BLURB["general_inquiry"])
    body = (
        f"Demo pack: {demo['slug']}\n"
        f"Expected intent: {intent_hint}\n"
        f"{'=' * 60}\n\n"
        f"How to run this case:\n"
        f"  1. From your email client signed in as rituraj@leewayhertz.com,\n"
        f"     compose a new message to the Keysight mailbox that the demo\n"
        f"     instance is polling (the one configured under\n"
        f"     Settings -> Integrations -> Email accounts).\n"
        f"  2. Paste the subject and body from email.txt below.\n"
        f"  3. Attach BOTH files in this folder:\n"
        f"       - {excel_name}\n"
        f"       - {pdf_name}\n"
        f"  4. Send the email.\n"
        f"  5. The IMAP poller picks it up within ten seconds. Watch the\n"
        f"     Dashboard inbox tile for the new row, then click into the\n"
        f"     Trace to follow Stage 1 to Stage 5.\n\n"
        f"What the case demonstrates:\n"
        f"{intent_blurb}\n\n"
        f"Subject (copy this verbatim):\n"
        f"  {demo['subject']}\n"
    )
    path.write_text(body, encoding="utf-8")
    return path


def main() -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    print(f"Writing demo packs under {OUTPUT_ROOT}")
    for demo in DEMOS:
        folder = OUTPUT_ROOT / demo["slug"]
        folder.mkdir(parents=True, exist_ok=True)
        write_email_txt(folder, demo)
        write_excel(folder, demo)
        write_pdf(folder, demo)
        write_readme(folder, demo)
        print(f"  [{demo['slug']}] email.txt + {demo['excel']['filename']} + {demo['pdf']['filename']}")
    # Top-level index file
    index = OUTPUT_ROOT / "INDEX.txt"
    lines = [
        "Leeway Hertz demo packs (mixed intents)",
        "=" * 60,
        "",
        "Sender to use in the From line: rituraj@leewayhertz.com",
        "  (added to Salesforce as a Contact under the Leeway Hertz",
        "   account, Contact Id 003dM00001xu0FlQAI)",
        "",
        f"{len(DEMOS)} demo cases:",
    ]
    for d in DEMOS:
        hint = d.get("intent_hint", "general_inquiry")
        lines.append(f"  - {d['slug']:34s} [{hint:15s}] {d['subject']}")
    lines.append("")
    lines.append("Intent mix this pack exercises:")
    lines.append("  general_inquiry  ->  01 lead time, 02 EOL roadmap")
    lines.append("  hold_release     ->  03 SO-LH-8214 payment cleared")
    lines.append("  po_intake        ->  04 PO-LH-2026-7700 (N9020B + N7045A)")
    lines.append("  quote_to_order   ->  05 QT-LH-9920 quote acceptance")
    lines.append("")
    lines.append("Each subfolder contains its own README.txt with step-by-step instructions.")
    index.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {index}")


if __name__ == "__main__":
    main()
